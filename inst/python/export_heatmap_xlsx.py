from __future__ import annotations
from pathlib import Path
from typing import Optional, Sequence
import pandas as pd
import openpyxl

def py_write_heatmap_xlsx(
    export_df,
    file: str,
    sheet: str = "Sheet1",
    expr_cols: Optional[Sequence[str]] = None,
    cluster_col_name: str = "cluster",
    replace_zero_annotation_with_blank: bool = True,
):
    """
    Write heatmap export table to Excel using pandas + xlsxwriter,
    safely handling non-UTF-8 characters.

    - Expression columns: auto-detect 'TPMmean' unless specified.
    - Annotation columns: all other columns.
    - TPMmean columns get row-wise 2-color scale.
    - Always freeze first row and first column.
    """

    df = pd.DataFrame(export_df).copy()

    # --- Safe UTF-8 conversion ---
    for col in df.select_dtypes(include=["object", "string"]).columns:
        df[col] = df[col].apply(
            lambda x: str(x).encode("utf-8", errors="replace").decode("utf-8")
        )

    # Ensure Gene_ID exists
    if "Gene_ID" not in df.columns:
        df.insert(0, "Gene_ID", df.index.astype(str))

    # Detect expression columns
    if expr_cols is None:
        expr_cols = [c for c in df.columns if "TPMmean" in str(c)]
    else:
        expr_cols = list(expr_cols)

    missing_expr = [c for c in expr_cols if c not in df.columns]
    if missing_expr:
        raise ValueError("These expr_cols are not in export_df: " + ", ".join(missing_expr))

    expr_set = set(expr_cols)
    annotation_cols = [c for c in df.columns if c not in expr_set]

    out_path = Path(file)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # --- Write with xlsxwriter ---
    with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet, na_rep="")
        workbook = writer.book
        worksheet = writer.sheets[sheet]

        n_rows, n_cols = df.shape

        # Formats
        numeric_center = workbook.add_format({"num_format": "0.00", "align": "center"})
        text_left = workbook.add_format({"num_format": "@", "align": "left"})
        rotated_header = workbook.add_format({"align": "center", "rotation": 90, "bold": True})

        # Column indices
        expr_idx = [df.columns.get_loc(c) for c in expr_cols]
        annotation_idx = [df.columns.get_loc(c) for c in annotation_cols]

        # Set column widths & formats
        for i, col in enumerate(df.columns):
            if col == "Gene_ID":
                worksheet.set_column(i, i, 22, text_left)
            elif i in expr_idx:
                worksheet.set_column(i, i, 7, numeric_center)
            else:
                worksheet.set_column(i, i, 12, text_left)

        # Rotate header
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, rotated_header)

        # --- Always freeze first row and first column ---
        worksheet.freeze_panes(1, 1)

        # TPMmean row-wise 2-color scale
        if expr_idx:
            expr_min = min(expr_idx)
            expr_max = max(expr_idx)
            for row_idx in range(1, n_rows + 1):
                worksheet.conditional_format(
                    row_idx,
                    expr_min,
                    row_idx,
                    expr_max,
                    {"type": "2_color_scale", "min_color": "#d4ffd1", "max_color": "#32cfff"},
                )

    # Replace 0 with blank in annotation columns
    if replace_zero_annotation_with_blank and annotation_cols:
        wb = openpyxl.load_workbook(out_path)
        ws = wb[sheet]
        ann_excel_cols = [df.columns.get_loc(c) + 1 for c in annotation_cols]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_num in ann_excel_cols:
                cell = row[col_num - 1]
                if cell.value == 0:
                    cell.value = None
        wb.save(out_path)

    return str(out_path)
